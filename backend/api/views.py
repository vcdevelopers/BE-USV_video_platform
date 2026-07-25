import json
import time
import datetime
import jwt
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import bcrypt
except ImportError:
    bcrypt = None

from .models import AdminUser, Language, Category, VideoSlot, Video, ViewEvent, SiteSetting
from .utils import verify_admin_token, detect_device_type, extract_ip_geo, save_uploaded_file, COUNTRY_NAME_MAP

def check_admin_password(password, stored_hash):
    if not stored_hash or not password:
        return False
    if bcrypt:
        try:
            return bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8'))
        except Exception:
            pass
    from django.contrib.auth.hashers import check_password
    try:
        if check_password(password, stored_hash):
            return True
    except Exception:
        pass
    return password == stored_hash


# -------------------------------------------------------------------
# PUBLIC ENDPOINTS
# -------------------------------------------------------------------

@csrf_exempt
@api_view(['GET'])
def get_languages(request):
    try:
        languages = list(Language.objects.filter(is_active=True).order_by('sort_order').values('code', 'name'))
        return JsonResponse({'languages': languages})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['GET'])
def get_site_settings(request):
    try:
        settings_dict = {}
        for s in SiteSetting.objects.all():
            try:
                settings_dict[s.key] = json.loads(s.value)
            except Exception:
                settings_dict[s.key] = s.value
        return JsonResponse({'settings': settings_dict})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['GET'])
def get_videos(request):
    try:
        target_lang = request.GET.get('lang', 'en').lower()

        categories = Category.objects.all().order_by('sort_order')
        slots = VideoSlot.objects.select_related('category').all().order_by('sort_order')
        active_videos = Video.objects.filter(is_active=True)

        video_map = {}
        for v in active_videos:
            if v.slot_id not in video_map:
                video_map[v.slot_id] = {}
            code = v.language_code.lower() if v.language_code else 'en'
            video_map[v.slot_id][code] = v

        categories_with_content = []
        for cat in categories:
            cat_slots = [s for s in slots if s.category_id == cat.id]
            items = []
            for slot in cat_slots:
                slot_vids = video_map.get(slot.id, {})
                matching_video = slot_vids.get(target_lang) or (list(slot_vids.values())[0] if slot_vids else None)

                items.append({
                    'id': slot.id,
                    'video_id': matching_video.id if matching_video else None,
                    'slug': slot.slug,
                    'title': slot.title,
                    'description': slot.description,
                    'is_intro': 1 if slot.is_intro else 0,
                    'video_url': matching_video.video_url if matching_video else None,
                    'thumbnail_url': matching_video.thumbnail_url if matching_video else None,
                    'target_area': matching_video.target_area if matching_video else None,
                    'language_code': matching_video.language_code if matching_video else 'en',
                    'has_video': bool(matching_video)
                })

            categories_with_content.append({
                'id': cat.id,
                'slug': cat.slug,
                'title': cat.title,
                'description': cat.description,
                'items': items
            })

        return JsonResponse({'categories': categories_with_content})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['GET'])
def play_video(request, video_id):
    try:
        admin_payload = verify_admin_token(request)
        is_admin = admin_payload is not None

        try:
            video = Video.objects.select_related('slot', 'slot__category').get(id=video_id)
        except Video.DoesNotExist:
            return JsonResponse({'error': 'Video guide not found.'}, status=404)

        # Inactive video enforcement
        if not video.is_active and not is_admin:
            return JsonResponse({'error': 'This video guide is currently inactive and unavailable.'}, status=403)

        lang_name = video.language_code
        try:
            lang_obj = Language.objects.get(code=video.language_code)
            lang_name = lang_obj.name
        except Language.DoesNotExist:
            pass

        video_detail = {
            'video_id': video.id,
            'slot_id': video.slot.id,
            'language_code': video.language_code,
            'video_url': video.video_url,
            'thumbnail_url': video.thumbnail_url,
            'target_area': video.target_area,
            'is_active': 1 if video.is_active else 0,
            'slot_title': video.slot.title,
            'slot_description': video.slot.description,
            'category_title': video.slot.category.title,
            'language_name': lang_name
        }

        return JsonResponse({'video': video_detail})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['POST'])
def log_view_event(request):
    try:
        data = request.data or {}
        language = data.get('language')
        if not language:
            return JsonResponse({'error': 'Language parameter is required'}, status=400)

        video_id = data.get('video_id')
        client_source = data.get('source')

        # SOURCE TRUST BOUNDARY ENFORCEMENT:
        # Must default to 'patient' server-side.
        # Only ever set to 'admin' when the request carries a valid authenticated admin JWT.
        admin_payload = verify_admin_token(request)
        if admin_payload and client_source == 'admin':
            event_source = 'admin'
        else:
            event_source = 'patient'

        device_type = detect_device_type(request)
        ip, country, city = extract_ip_geo(request)

        video_obj = None
        if video_id:
            try:
                video_obj = Video.objects.get(id=video_id)
            except Video.DoesNotExist:
                pass

        event = ViewEvent.objects.create(
            language=language,
            device_type=device_type,
            ip_address=ip,
            country=country,
            city=city,
            video=video_obj,
            source=event_source
        )

        return JsonResponse({
            'success': True,
            'logged': {
                'id': event.id,
                'language': language,
                'video_id': video_id,
                'deviceType': device_type,
                'country': country,
                'city': city,
                'source': event_source
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# -------------------------------------------------------------------
# ADMIN AUTH ENDPOINTS
# -------------------------------------------------------------------

@csrf_exempt
@api_view(['POST'])
def admin_login(request):
    try:
        data = request.data or {}
        email = data.get('email')
        password = data.get('password')

        if not email or not password:
            return JsonResponse({'error': 'Email and password are required'}, status=400)

        try:
            admin = AdminUser.objects.get(email=email)
        except AdminUser.DoesNotExist:
            return JsonResponse({'error': 'Invalid email or password'}, status=401)

        if not check_admin_password(password, admin.password_hash):
            return JsonResponse({'error': 'Invalid email or password'}, status=401)

        payload = {
            'id': admin.id,
            'email': admin.email,
            'exp': int(time.time()) + (24 * 3600) # 24 hour expiry
        }
        token = jwt.encode(payload, settings.JWT_SECRET, algorithm='HS256')

        return JsonResponse({
            'token': token,
            'admin': {'id': admin.id, 'email': admin.email}
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['GET'])
def admin_me(request):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)
    return JsonResponse({'admin': admin_payload})


@csrf_exempt
@api_view(['POST'])
def update_site_settings(request):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    try:
        data = request.data or {}
        disclaimer = data.get('disclaimer')
        important_tips = data.get('important_tips')

        if disclaimer is not None:
            SiteSetting.objects.update_or_create(key='disclaimer', defaults={'value': disclaimer})

        if important_tips is not None:
            tips_val = json.dumps(important_tips) if not isinstance(important_tips, str) else important_tips
            SiteSetting.objects.update_or_create(key='important_tips', defaults={'value': tips_val})

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# -------------------------------------------------------------------
# ADMIN VIDEO MANAGEMENT ENDPOINTS
# -------------------------------------------------------------------

@csrf_exempt
@api_view(['GET', 'POST'])
def admin_videos(request):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    if request.method == 'GET':
        try:
            slots_qs = VideoSlot.objects.select_related('category').all().order_by('category__sort_order', 'sort_order')
            slots = [{
                'slot_id': s.id,
                'slot_title': s.title,
                'slot_description': s.description,
                'category_id': s.category.id,
                'category_title': s.category.title,
                'is_intro': 1 if s.is_intro else 0
            } for s in slots_qs]

            videos_qs = Video.objects.select_related('slot', 'slot__category').annotate(
                views=Count('view_events', filter=Q(view_events__source='patient'))
            ).order_by('-id')

            videos = [{
                'id': v.id,
                'slot_id': v.slot.id,
                'video_url': v.video_url,
                'thumbnail_url': v.thumbnail_url,
                'target_area': v.target_area,
                'is_active': 1 if v.is_active else 0,
                'created_at': v.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'slot_title': v.slot.title,
                'slot_description': v.slot.description,
                'category_id': v.slot.category.id,
                'category_title': v.slot.category.title,
                'views': v.views
            } for v in videos_qs]

            languages = list(Language.objects.all().order_by('sort_order').values('code', 'name'))
            categories = list(Category.objects.all().order_by('sort_order').values('id', 'title'))

            return JsonResponse({'slots': slots, 'videos': videos, 'languages': languages, 'categories': categories})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    elif request.method == 'POST':
        try:
            data = request.POST
            files = request.FILES

            slot_id = data.get('slot_id')
            category_id = data.get('category_id')
            new_cat_title = data.get('new_category_title')
            new_cat_desc = data.get('new_category_description', '')
            new_slot_title = data.get('new_slot_title') or data.get('slot_title')
            new_slot_desc = data.get('new_slot_description') or data.get('slot_description', '')
            lang_code_input = data.get('language_code')
            lang_name_input = data.get('language_name')
            target_area = data.get('target_area', '')
            is_active_input = data.get('is_active')
            video_url_custom = data.get('video_url_custom', '')
            thumb_url_custom = data.get('thumbnail_url_custom', '')

            active_flag = False if is_active_input in ('false', 0, '0') else True

            target_cat_id = int(category_id) if category_id and category_id != 'new' else None
            if (not target_cat_id or category_id == 'new') and new_cat_title:
                cat_slug = new_cat_title.lower().replace(' ', '-').replace('/', '-') + '-' + str(int(time.time()))
                new_cat = Category.objects.create(
                    slug=cat_slug,
                    title=new_cat_title,
                    description=new_cat_desc,
                    sort_order=99
                )
                target_cat_id = new_cat.id

            if lang_code_input and lang_name_input:
                Language.objects.get_or_create(
                    code=lang_code_input.lower(),
                    defaults={'name': lang_name_input, 'is_active': True, 'sort_order': 99}
                )
            lang_code = lang_code_input.lower() if lang_code_input else 'en'

            target_slot_id = int(slot_id) if slot_id and slot_id != 'new' else None
            if not target_slot_id or slot_id == 'new':
                if not target_cat_id:
                    return JsonResponse({'error': 'Please select an existing category or create a custom category.'}, status=400)
                if not new_slot_title:
                    return JsonResponse({'error': 'Please enter a title for the new video slot.'}, status=400)

                slot_slug = new_slot_title.lower().replace(' ', '-').replace('/', '-') + '-' + str(int(time.time()))
                cat_obj = Category.objects.get(id=target_cat_id)
                new_slot = VideoSlot.objects.create(
                    category=cat_obj,
                    slug=slot_slug,
                    title=new_slot_title,
                    description=new_slot_desc,
                    is_intro=False,
                    sort_order=99
                )
                target_slot_id = new_slot.id

            final_video_url = video_url_custom
            if 'video_file' in files:
                final_video_url = save_uploaded_file(files['video_file'], 'video')

            final_thumb_url = thumb_url_custom
            if 'thumbnail_file' in files:
                final_thumb_url = save_uploaded_file(files['thumbnail_file'], 'thumbnail')

            if not final_video_url and 'video_file' not in files:
                existing_vid = Video.objects.filter(slot_id=target_slot_id, language_code=lang_code).first()
                if not existing_vid or not existing_vid.video_url:
                    return JsonResponse({'error': 'Please upload a video file or provide a valid video URL.'}, status=400)

            # UPDATE-IN-PLACE UNICITY RULE: (slot_id, language_code)
            existing = Video.objects.filter(slot_id=target_slot_id, language_code=lang_code).first()

            if existing:
                if final_video_url:
                    existing.video_url = final_video_url
                if final_thumb_url or 'thumbnail_file' in files:
                    existing.thumbnail_url = final_thumb_url
                existing.target_area = target_area
                existing.is_active = active_flag
                existing.save()
                return JsonResponse({'success': True, 'video_id': existing.id, 'slot_id': target_slot_id, 'language_code': lang_code, 'action': 'updated'})
            else:
                slot_obj = VideoSlot.objects.get(id=target_slot_id)
                new_vid = Video.objects.create(
                    slot=slot_obj,
                    language_code=lang_code,
                    video_url=final_video_url,
                    thumbnail_url=final_thumb_url,
                    target_area=target_area,
                    is_active=active_flag
                )
                return JsonResponse({'success': True, 'video_id': new_vid.id, 'slot_id': target_slot_id, 'language_code': lang_code, 'action': 'created'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)



@csrf_exempt
@api_view(['POST'])
def admin_create_or_update_video(request):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    try:
        data = request.POST
        files = request.FILES

        slot_id = data.get('slot_id')
        category_id = data.get('category_id')
        new_cat_title = data.get('new_category_title')
        new_cat_desc = data.get('new_category_description', '')
        new_slot_title = data.get('new_slot_title') or data.get('slot_title')
        new_slot_desc = data.get('new_slot_description') or data.get('slot_description', '')
        lang_code_input = data.get('language_code')
        lang_name_input = data.get('language_name')
        target_area = data.get('target_area', '')
        is_active_input = data.get('is_active')
        video_url_custom = data.get('video_url_custom', '')
        thumb_url_custom = data.get('thumbnail_url_custom', '')

        active_flag = False if is_active_input in ('false', 0, '0') else True

        # Handle category creation if new
        target_cat_id = int(category_id) if category_id and category_id != 'new' else None
        if (not target_cat_id or category_id == 'new') and new_cat_title:
            cat_slug = new_cat_title.lower().replace(' ', '-').replace('/', '-') + '-' + str(int(time.time()))
            new_cat = Category.objects.create(
                slug=cat_slug,
                title=new_cat_title,
                description=new_cat_desc,
                sort_order=99
            )
            target_cat_id = new_cat.id

        # Language registration
        if lang_code_input and lang_name_input:
            Language.objects.get_or_create(
                code=lang_code_input.lower(),
                defaults={'name': lang_name_input, 'is_active': True, 'sort_order': 99}
            )
        lang_code = lang_code_input.lower() if lang_code_input else 'en'

        # Handle slot creation if new
        target_slot_id = int(slot_id) if slot_id and slot_id != 'new' else None
        if not target_slot_id or slot_id == 'new':
            if not target_cat_id:
                return JsonResponse({'error': 'Please select an existing category or create a custom category.'}, status=400)
            if not new_slot_title:
                return JsonResponse({'error': 'Please enter a title for the new video slot.'}, status=400)

            slot_slug = new_slot_title.lower().replace(' ', '-').replace('/', '-') + '-' + str(int(time.time()))
            cat_obj = Category.objects.get(id=target_cat_id)
            new_slot = VideoSlot.objects.create(
                category=cat_obj,
                slug=slot_slug,
                title=new_slot_title,
                description=new_slot_desc,
                is_intro=False,
                sort_order=99
            )
            target_slot_id = new_slot.id

        # File processing & server-side validation
        final_video_url = video_url_custom
        if 'video_file' in files:
            final_video_url = save_uploaded_file(files['video_file'], 'video')

        final_thumb_url = thumb_url_custom
        if 'thumbnail_file' in files:
            final_thumb_url = save_uploaded_file(files['thumbnail_file'], 'thumbnail')

        if not final_video_url and 'video_file' not in files:
            # If updating existing, check if video_url already exists
            existing_vid = Video.objects.filter(slot_id=target_slot_id, language_code=lang_code).first()
            if not existing_vid or not existing_vid.video_url:
                return JsonResponse({'error': 'Please upload a video file or provide a valid video URL.'}, status=400)

        # UPDATE-IN-PLACE UNICITY RULE: (slot_id, language_code)
        existing = Video.objects.filter(slot_id=target_slot_id, language_code=lang_code).first()

        if existing:
            if final_video_url:
                existing.video_url = final_video_url
            if final_thumb_url or 'thumbnail_file' in files:
                existing.thumbnail_url = final_thumb_url
            existing.target_area = target_area
            existing.is_active = active_flag
            existing.save()
            return JsonResponse({'success': True, 'video_id': existing.id, 'slot_id': target_slot_id, 'language_code': lang_code, 'action': 'updated'})
        else:
            slot_obj = VideoSlot.objects.get(id=target_slot_id)
            new_vid = Video.objects.create(
                slot=slot_obj,
                language_code=lang_code,
                video_url=final_video_url,
                thumbnail_url=final_thumb_url,
                target_area=target_area,
                is_active=active_flag
            )
            return JsonResponse({'success': True, 'video_id': new_vid.id, 'slot_id': target_slot_id, 'language_code': lang_code, 'action': 'created'})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['PUT', 'DELETE'])
def admin_edit_video(request, video_id):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    if request.method == 'DELETE':
        try:
            try:
                video = Video.objects.get(id=video_id)
                video.delete()
                return JsonResponse({'success': True})
            except Video.DoesNotExist:
                return JsonResponse({'error': 'Video not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    try:
        try:
            video = Video.objects.select_related('slot').get(id=video_id)
        except Video.DoesNotExist:
            return JsonResponse({'error': 'Video entry not found'}, status=404)

        data = request.POST
        files = request.FILES

        slot_title = data.get('slot_title')
        slot_description = data.get('slot_description')
        target_area = data.get('target_area')
        is_active_input = data.get('is_active')
        video_url_custom = data.get('video_url_custom')
        thumb_url_custom = data.get('thumbnail_url_custom')

        if slot_title:
            video.slot.title = slot_title
            if slot_description is not None:
                video.slot.description = slot_description
            video.slot.save()

        if 'video_file' in files:
            video.video_url = save_uploaded_file(files['video_file'], 'video')
        elif video_url_custom:
            video.video_url = video_url_custom

        if 'thumbnail_file' in files:
            video.thumbnail_url = save_uploaded_file(files['thumbnail_file'], 'thumbnail')
        elif thumb_url_custom is not None:
            video.thumbnail_url = thumb_url_custom

        if target_area is not None:
            video.target_area = target_area

        if is_active_input is not None:
            video.is_active = False if is_active_input in ('false', 0, '0') else True

        video.save()

        return JsonResponse({'success': True, 'video_id': video.id})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)




@csrf_exempt
@api_view(['PUT'])
def admin_toggle_video(request, video_id):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    try:
        try:
            video = Video.objects.get(id=video_id)
        except Video.DoesNotExist:
            return JsonResponse({'error': 'Video not found'}, status=404)

        video.is_active = not video.is_active
        video.save()

        return JsonResponse({'success': True, 'is_active': 1 if video.is_active else 0})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['DELETE'])
def admin_delete_video(request, video_id):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    try:
        try:
            video = Video.objects.get(id=video_id)
            video.delete()
            return JsonResponse({'success': True})
        except Video.DoesNotExist:
            return JsonResponse({'error': 'Video not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['POST'])
def admin_add_language(request):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    try:
        data = request.data or {}
        code = data.get('code')
        name = data.get('name')

        if not code or not name:
            return JsonResponse({'error': 'Code and Name are required'}, status=400)

        Language.objects.get_or_create(code=code.lower(), defaults={'name': name, 'is_active': True, 'sort_order': 99})

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# -------------------------------------------------------------------
# ADMIN ANALYTICS ENDPOINTS
# -------------------------------------------------------------------

@csrf_exempt
@api_view(['GET'])
def admin_analytics_summary(request):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    try:
        patient_events = ViewEvent.objects.filter(Q(source='patient') | Q(source__isnull=True), video__isnull=False)

        total_views = patient_events.count()

        lang_stats_qs = patient_events.values('language').annotate(views=Count('id')).order_by('-views')
        lang_map = {l.code: l.name for l in Language.objects.all()}
        lang_stats = [{
            'code': item['language'],
            'name': lang_map.get(item['language'], item['language']),
            'views': item['views']
        } for item in lang_stats_qs]

        per_video_qs = Video.objects.select_related('slot', 'slot__category').annotate(
            views=Count('view_events', filter=Q(view_events__source='patient') | Q(view_events__source__isnull=True))
        ).order_by('-views')

        per_video_stats = [{
            'video_id': v.id,
            'slot_title': v.slot.title,
            'category_title': v.slot.category.title,
            'views': v.views
        } for v in per_video_qs]

        highest_accessed = per_video_stats[0] if len(per_video_stats) > 0 else {'slot_title': 'None', 'views': 0}

        device_qs = patient_events.values('device_type').annotate(count=Count('id'))
        device_stats = [{
            'device': item['device_type'] or 'Desktop',
            'count': item['count']
        } for item in device_qs]

        cat_qs = Category.objects.annotate(
            views=Count('slots__videos__view_events', filter=Q(slots__videos__view_events__source='patient') | Q(slots__videos__view_events__source__isnull=True)),
            video_count=Count('slots__videos', distinct=True)
        ).order_by('-views')

        category_stats = [{
            'category_title': c.title,
            'views': c.views,
            'video_count': c.video_count
        } for c in cat_qs]

        return JsonResponse({
            'summary': {
                'totalViews': total_views,
                'highestAccessedVideo': highest_accessed,
                'totalVideos': len(per_video_stats)
            },
            'videoAnalytics': per_video_stats,
            'deviceAnalytics': device_stats,
            'categoryAnalytics': category_stats
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@api_view(['GET'])
def admin_analytics_export(request):
    admin_payload = verify_admin_token(request)
    if not admin_payload:
        return JsonResponse({'error': 'Unauthorized. Valid admin token required.'}, status=401)

    try:
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Analytics Summary'

        ws_summary.append(['Metric', 'Value'])
        total_views = ViewEvent.objects.filter(Q(source='patient') | Q(source__isnull=True), video__isnull=False).count()
        distinct_langs = ViewEvent.objects.filter(Q(source='patient') | Q(source__isnull=True), video__isnull=False).values('language').distinct().count()

        ws_summary.append(['Total Patient Views', total_views])
        ws_summary.append(['Distinct Languages Accessed', distinct_langs])
        ws_summary.append(['Report Generated At', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

        ws_lang = wb.create_sheet(title='Language Breakdown')
        ws_lang.append(['Language Code', 'Language Name', 'Total Patient Views'])

        lang_map = {l.code: l.name for l in Language.objects.all()}
        lang_qs = ViewEvent.objects.filter(Q(source='patient') | Q(source__isnull=True), video__isnull=False).values('language').annotate(views=Count('id')).order_by('-views')
        for item in lang_qs:
            ws_lang.append([item['language'], lang_map.get(item['language'], item['language']), item['views']])

        ws_video = wb.create_sheet(title='Per-Video Breakdown')
        ws_video.append(['Video ID', 'Exercise Slot', 'Category', 'Total Views'])

        vid_qs = Video.objects.select_related('slot', 'slot__category').annotate(
            views=Count('view_events', filter=Q(view_events__source='patient') | Q(view_events__source__isnull=True))
        ).order_by('-views')
        for v in vid_qs:
            ws_video.append([v.id, v.slot.title, v.slot.category.title, v.views])

        ws_device = wb.create_sheet(title='Device Breakdown')
        ws_device.append(['Device Type', 'Total Views'])
        dev_qs = ViewEvent.objects.filter(Q(source='patient') | Q(source__isnull=True)).values('device_type').annotate(count=Count('id'))
        for d in dev_qs:
            ws_device.append([d['device_type'] or 'desktop', d['count']])

        ws_raw = wb.create_sheet(title='Raw View Logs')
        ws_raw.append(['ID', 'Timestamp', 'Language', 'Video ID', 'Source', 'Device Type', 'IP Address', 'Country', 'City'])

        logs = ViewEvent.objects.all().order_by('-id')[:1000]
        for l in logs:
            ws_raw.append([
                l.id,
                l.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                l.language,
                l.video.id if l.video else None,
                l.source,
                l.device_type,
                l.ip_address,
                l.country,
                l.city
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Exercise_Video_Analytics_{int(time.time())}.xlsx"'

        wb.save(response)
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
